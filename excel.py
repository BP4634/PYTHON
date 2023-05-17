import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path

path = Path()




def workbook(filename):
    wb = xl.load_workbook(filename)

    sheet = wb["Hoja1"]


    for row in range(1, 11):
        for column in range(1, 11):
            
            cell = sheet.cell(row, column)
            new_value = sheet.cell(row, column)
            new_value.value = "HELLO"



    #values = Reference(sheet,
            #min_row=2,
            #max_row=sheet.max_row,
            #min_col=4,
            #max_col=4)

    #chart = BarChart()
    #chart.add_data(values)

    #sheet.add_chart(chart, 'e2')





    wb.save(filename)


for file in path.glob("*.xlsx"):
    
    workbook(file)


